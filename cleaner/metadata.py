"""
元数据归一 — 统一提取各类文件的元数据
"""
import json
from pathlib import Path
from datetime import datetime

from cleaner.state import upsert_metadata, log


def extract(file_path: str, file_type: str, file_hash: str, conn) -> dict:
    """
    提取元数据并写入 file_metadata 表
    返回提取的元数据字典
    """
    path = Path(file_path)
    meta = {
        "file_size": path.stat().st_size,
        "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
    }

    if file_type == "docx":
        _extract_docx(path, meta)
    elif file_type == "xlsx":
        _extract_xlsx(path, meta)
    elif file_type == "pptx":
        _extract_pptx(path, meta)
    elif file_type == "pdf":
        _extract_pdf(path, meta)
    elif file_type == "image":
        _extract_image(path, meta)
    elif file_type in ("video", "audio"):
        _extract_media(path, file_type, meta)

    # 写入数据库
    upsert_metadata(conn, file_hash, **meta)
    log(conn, file_hash, "metadata", "extracted", json.dumps(meta, ensure_ascii=False, default=str))

    return meta


def _extract_docx(path, meta):
    try:
        from docx import Document
        doc = Document(str(path))
        props = doc.core_properties
        meta["title"] = props.title or None
        meta["author"] = props.author or None
        meta["page_count"] = len(doc.paragraphs) // 30 or None  # 粗略估算
    except Exception:
        pass


def _extract_xlsx(path, meta):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        meta["title"] = wb.properties.title or None
        meta["author"] = wb.properties.creator or None
        meta["sheet_count"] = len(wb.sheetnames)
        wb.close()
    except Exception:
        pass


def _extract_pptx(path, meta):
    try:
        from pptx import Presentation
        prs = Presentation(str(path))
        props = prs.core_properties
        meta["title"] = props.title or None
        meta["author"] = props.author or None
        meta["slide_count"] = len(prs.slides)
    except Exception:
        pass


def _extract_pdf(path, meta):
    try:
        import pymupdf
        doc = pymupdf.open(str(path))
        pdf_meta = doc.metadata
        meta["title"] = pdf_meta.get("title") or None
        meta["author"] = pdf_meta.get("author") or None
        meta["page_count"] = len(doc)
        doc.close()
    except Exception:
        pass


def _extract_image(path, meta):
    try:
        from PIL import Image
        from PIL.ExifTags import TAGS
        with Image.open(path) as img:
            meta["resolution"] = f"{img.width}x{img.height}"
            exif = img._getexif()
            if exif:
                for tag_id, value in exif.items():
                    tag = TAGS.get(tag_id, tag_id)
                    if tag == "DateTimeOriginal":
                        meta["created_at"] = value
    except Exception:
        pass


def _extract_media(path, file_type, meta):
    try:
        import subprocess
        result = subprocess.run(
            ["ffprobe", "-v", "quiet", "-print_format", "json", "-show_format", "-show_streams", str(path)],
            capture_output=True, text=True, timeout=10,
        )
        if result.returncode == 0:
            data = json.loads(result.stdout)
            fmt = data.get("format", {})
            meta["duration_sec"] = float(fmt.get("duration", 0)) or None

            for stream in data.get("streams", []):
                if stream.get("codec_type") == "video":
                    meta["resolution"] = f"{stream.get('width')}x{stream.get('height')}"
                    break
    except Exception:
        pass
