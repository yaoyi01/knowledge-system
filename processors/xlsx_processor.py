"""XLSX 处理器"""
from pathlib import Path
from cleaner.state import upsert_metadata, log

BASE_DIR = Path.home() / "Documents" / "knowledge-system"


def process_xlsx(file_hash: str, vault_path: str, conn) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError:
        log(conn, file_hash, "processor", "error", "openpyxl 未安装")
        return None

    wb = load_workbook(vault_path, read_only=True, data_only=True)
    lines = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n## Sheet: {sheet_name}\n")
        for row in ws.iter_rows(values_only=True):
            cleaned = [str(c).strip() if c is not None else "" for c in row]
            if any(cleaned):
                lines.append("\t".join(cleaned))

    wb.close()
    text = "\n".join(lines)
    text_path = BASE_DIR / "processed" / "text" / f"{file_hash}.txt"
    text_path.write_text(text, encoding="utf-8")

    meta = {"sheet_count": len(wb.sheetnames)}
    upsert_metadata(conn, file_hash, **meta)

    log(conn, file_hash, "processor", "done", f"xlsx→text: {len(text)} chars, {len(wb.sheetnames)} sheets")
    return {"text_path": str(text_path), "image_paths": [], "meta": meta}
